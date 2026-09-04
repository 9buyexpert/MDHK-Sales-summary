#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Monopoly Dreams POS 自動化報表腳本
- 登入 Monopoly Dreams POS (https://www.boxasone.net:2888)
- 抓取 Kiosk VM 1082 及 1083 之 Product Sold Summary 與 Discount Applied Summary
- 產生 "Month year.xlsx"
- 上傳至 Google Drive 2026 資料夾
"""

import os
import sys
import calendar
import argparse
from datetime import datetime, date
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from playwright.sync_api import sync_playwright

# Google Drive API
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# ----------------- 設定區 -----------------
POS_LOGIN_URL = "https://www.boxasone.net:2888/Login.aspx"
POS_REPORT_URL = "https://www.boxasone.net:2888/Report.aspx"
POS_USER = "admin"
POS_PASS = "daydreaming"

TARGET_KIOSKS = ["1082", "1083"]
DRIVE_FOLDER_ID = "1XFGCC5XwZgIXYiK-4dDNW_WjwdjZRtBr"
SERVICE_ACCOUNT_FILE = "service_account.json"
# ------------------------------------------

def get_month_date_range(year: int, month: int):
    """取得指定月份的第一天與最後一天字串 (YYYY-MM-DD)"""
    first_day = date(year, month, 1).strftime("%Y-%m-%d")
    last_day_num = calendar.monthrange(year, month)[1]
    last_day = date(year, month, last_day_num).strftime("%Y-%m-%d")
    month_name = calendar.month_name[month]
    return first_day, last_day, f"{month_name} {year}"

def scrape_pos_data(start_date: str, end_date: str, kiosk_ids: list):
    """
    使用 Playwright 登入並獲取指定日期與機台的報表 HTML
    """
    results = {}
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(ignore_https_errors=True)
        page = context.new_page()

        print(f"[*] 正在連線至 POS 登入頁面: {POS_LOGIN_URL}")
        page.goto(POS_LOGIN_URL, timeout=60000)

        # 登入處理
        try:
            if page.locator('input[name="txtUserName"], input[id*="UserName"], input[id*="txtUser"]').count() > 0:
                page.fill('input[name="txtUserName"], input[id*="UserName"], input[id*="txtUser"]', POS_USER)
                page.fill('input[name="txtPassword"], input[id*="Password"], input[id*="txtPass"]', POS_PASS)
                page.click('input[type="submit"], button[type="submit"], input[id*="btnLogin"]')
                page.wait_for_load_state("networkidle")
                print("[+] 登入成功")
        except Exception as e:
            print(f"[!] 登入檢查提示: {e}")

        # 前往報表頁面
        page.goto(POS_REPORT_URL, timeout=60000)
        page.wait_for_load_state("networkidle")

        for kiosk in kiosk_ids:
            print(f"[*] 正在查詢機台 VM {kiosk} ({start_date} 至 {end_date})...")
            
            # 填寫日期
            # 依據 ASP.NET 控制項名稱進行定位
            if page.locator('input[id*="txtStartDate"], input[name*="txtStartDate"]').count() > 0:
                page.fill('input[id*="txtStartDate"], input[name*="txtStartDate"]', start_date)
            if page.locator('input[id*="txtEndDate"], input[name*="txtEndDate"]').count() > 0:
                page.fill('input[id*="txtEndDate"], input[name*="txtEndDate"]', end_date)

            # 勾選 Show Items and Discount Summary
            chk_selector = 'input[type="checkbox"][id*="chkShowItem"], input[type="checkbox"][id*="Discount"]'
            checkboxes = page.locator(chk_selector)
            for i in range(checkboxes.count()):
                chk = checkboxes.nth(i)
                if not chk.is_checked():
                    chk.check()

            # 選擇機台 VM
            select_locator = page.locator('select[id*="ddlKiosk"], select[id*="Kiosk"], select[id*="Machine"]')
            if select_locator.count() > 0:
                try:
                    select_locator.first.select_option(value=kiosk)
                except Exception:
                    select_locator.first.select_option(label=f"Kiosk {kiosk}")

            # 點擊查詢 / 提交
            submit_btn = page.locator('input[type="submit"][value*="Search"], input[type="submit"][value*="Submit"], input[id*="btnSearch"]')
            if submit_btn.count() > 0:
                submit_btn.first.click()
                page.wait_for_load_state("networkidle")
            
            # 取得結果 HTML
            results[kiosk] = page.content()
            print(f"[+] 機台 VM {kiosk} 查詢完成")

        browser.close()
    return results

def parse_report_html(html_content: str):
    """
    解析 Product Sold Summary 與 Discount Applied Summary 表格
    """
    soup = BeautifulSoup(html_content, "html.parser")
    product_rows = []
    discount_rows = []

    # 尋找所有表格進行識別
    tables = soup.find_all("table")
    for tbl in tables:
        tbl_text = tbl.get_text()
        if "Product Sold Summary" in tbl_text or "Ean Code" in tbl_text:
            for tr in tbl.find_all("tr"):
                cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
                if cells and not any("Product Sold Summary" in c for c in cells):
                    product_rows.append(cells)
        elif "Discount Applied" in tbl_text or "Discount Summary" in tbl_text:
            for tr in tbl.find_all("tr"):
                cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
                if cells:
                    discount_rows.append(cells)

    return product_rows, discount_rows

def build_excel_file(month_label: str, all_kiosk_data: dict, output_filepath: str):
    """
    參照 April 2026.xlsx 結構建立 Excel 檔
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Summary"

    # 預設商品欄位標題
    headers = [
        "Category", "Ean Code", "Description", "Qty", "Unit Price",
        "2P Shares %", "2P Shares", "FMP Shares %", "FMP Shares", "Amount", "COST", "Amount cost"
    ]
    
    # 寫入第一行：月份 + 欄位
    ws.append([f"{month_label} Category"] + headers[1:])
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)

    # 寫入各機台銷售商品資料
    for kiosk, (products, _) in all_kiosk_data.items():
        for r in products:
            # 避免重複寫入標題行
            if len(r) >= 4 and not any("Category" in str(x) or "Ean Code" in str(x) for x in r):
                ws.append(r)

    # 空行分隔
    ws.append([])
    ws.append(["Discount Applied / Ticket Summary"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["", "Description", "Qty", "Unit Price", "Amount"])

    # 寫入折扣與票券摘要
    for kiosk, (_, discounts) in all_kiosk_data.items():
        for r in discounts:
            if len(r) >= 3 and not any("Description" in str(x) for x in r):
                ws.append([""] + r)

    wb.save(output_filepath)
    print(f"[+] Excel 檔案已成功生成: {output_filepath}")

def upload_to_drive(local_filepath: str, folder_id: str):
    """
    將產生的 Excel 檔案上傳到指定 Google Drive 資料夾
    """
    if not os.path.exists(SERVICE_ACCOUNT_FILE):
        print(f"[!] 未找到 {SERVICE_ACCOUNT_FILE}，請確認服務帳號憑證已設定。已儲存於本機：{local_filepath}")
        return None

    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE,
        scopes=['https://www.googleapis.com/auth/drive']
    )
    service = build('drive', 'v3', credentials=creds)

    filename = os.path.basename(local_filepath)
    file_metadata = {
        'name': filename,
        'parents': [folder_id]
    }
    media = MediaFileUpload(
        local_filepath,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        resumable=True
    )

    file = service.files().create(
        body=file_metadata,
        media_body=media,
        fields='id, webViewLink'
    ).execute()

    print(f"[+] 檔案成功上傳至 Google Drive: {file.get('webViewLink')}")
    return file.get('id')

def main():
    parser = argparse.ArgumentParser(description="Monopoly Dreams POS Monthly Report Generator")
    parser.add_argument("--year", type=int, help="年份 (例: 2026)")
    parser.add_argument("--month", type=int, help="月份 (1-12)")
    args = parser.parse_args()

    # 若未指定年月，預設抓取上個月（適合每月 1 號執行）
    today = date.today()
    if args.year and args.month:
        target_year, target_month = args.year, args.month
    else:
        if today.month == 1:
            target_year, target_month = today.year - 1, 12
        else:
            target_year, target_month = today.year, today.month - 1

    start_date, end_date, month_label = get_month_date_range(target_year, target_month)
    output_filename = f"{month_label}.xlsx"

    print(f"=== 開始處理報表: {month_label} ({start_date} ~ {end_date}) ===")

    # 1. 爬取資料
    raw_html_dict = scrape_pos_data(start_date, end_date, TARGET_KIOSKS)

    # 2. 解析資料
    parsed_data = {}
    for kiosk, html in raw_html_dict.items():
        products, discounts = parse_report_html(html)
        parsed_data[kiosk] = (products, discounts)

    # 3. 輸出 Excel
    build_excel_file(month_label, parsed_data, output_filename)

    # 4. 上傳 Google Drive
    upload_to_drive(output_filename, DRIVE_FOLDER_ID)
    print("=== 任務執行完成 ===")

if __name__ == "__main__":
    main()